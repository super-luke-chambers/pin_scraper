# INPUTS
from datetime import datetime
RUN_ID = datetime.now().strftime("%d_%b_%Y").upper() + "_Run1"
INPUT_FILE = "pinnacle_league_inputs_web.xlsx"
INPUT_SHEET = "Sheet1"
DATABASE_NAME = "PinnacleOddsDatabase.db"
PATH_TO_BROWSER = "" # shouldn't matter anymore!
INPUT_MAX_FUTURE_DAYS = 7 # to avoid scraping matches 1 week from now which are displayed
# END OF INPUTS


# install: selenium, lxml, openpyxl
from selenium import webdriver  #install selenium
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from lxml import html
from openpyxl import Workbook, load_workbook
import os
import sys
import bz2
import json
import pickle
import pprint
import threading
import time
from datetime import datetime
import sqlite3
from urllib.parse import urljoin



class PinnacleScraper:
    def __init__(self, input_runid, input_filename, input_sheetname, input_databasename, input_browserpath, input_maxfuturedays, mode='scrape'):
        ## check if inputs are good
        self.inputs_are_good = True
        self.is_interrupted = False
        self.TIME_PAUSE = 1.0 # pause for xpath

        input_checks = [self.check_input('RUN_ID', 'str', input_runid),
                        self.check_input('INPUT_FILE', 'str', input_filename),
                        self.check_input('INPUT_SHEET', 'str', input_sheetname),
                        self.check_input('DATABASE_NAME', 'str', input_databasename),
                        self.check_input('PATH_TO_BROWSER', 'str', input_browserpath),
                        self.check_input('INPUT_MAX_FUTURE_DAYS', 'positive_int', input_maxfuturedays)
                        ]
        if False in input_checks:
            print("Bad inputs, quit!")
            self.inputs_are_good = False
            return

        ## if still here, set inputs
        self.run_id = input_runid
        self.input_file = input_filename
        self.input_sheet = input_sheetname
        self.database_name = input_databasename
        self.PATH_TO_BROWSER = input_browserpath
        self.max_future_days = input_maxfuturedays

        ## read inputs
        self.leagues_to_scrape = self.read_inputs()

        ## create database
        if not os.path.exists(self.database_name):
            print("Creating a new database...")
        else:
            print("Database already exists!")
        self.db_conn = sqlite3.connect(self.database_name, check_same_thread=False)
        self.db_cursor = self.db_conn.cursor()
        self.db_cursor.execute("CREATE TABLE IF NOT EXISTS ScrapedRuns (run_id TEXT NOT NULL, league_url TEXT NOT NULL, number_of_games INTEGER, PRIMARY KEY(run_id, league_url))")
        self.db_cursor.execute("""CREATE TABLE IF NOT EXISTS Matches (game_url TEXT NOT NULL, run_id TEXT NOT NULL, home_team TEXT, away_team TEXT, match_time TEXT, match_timestamp REAL, league_name TEXT, league_url TEXT,
                                    odds_html BLOB, time_of_scraping TEXT, time_of_scraping_stamp REAL, PRIMARY KEY(game_url, run_id))""")


        ## start chrome
        if mode == 'scrape':
            self.driver = self.start_driver_normal()
        elif mode == 'headless':
            self.driver = self.start_driver_headless()
        return

    def read_inputs(self):
        try:
            items_to_return = []
            input_wb = load_workbook(self.input_file)
            input_ws = input_wb[self.input_sheet]
            for input_row_number in range(2, input_ws.max_row+1):
                league_url = input_ws.cell(row=input_row_number, column=1).value
                league_name = input_ws.cell(row=input_row_number, column=2).value

                if type(league_name) == str and type(league_url) == str:
                    if "//www.pinnacle.com/" in league_url:
                        # good to add
                        item_to_add = {"url":league_url, "league":league_name}
                        if item_to_add not in items_to_return:
                            items_to_return.append(item_to_add)
                else:
                    print("Wrong input types in row", input_row_number)
                            
            return items_to_return
        except:
            print("An exception while reading inputs - make sure input filename and sheetname are correct!")
            return []
        
    def check_input(self, input_name, input_type, input_value):
        input_is_good = True
        if input_type == 'str':
            if type(input_value) != str:
                input_is_good = False
                print(input_name + " should be a string!")
        elif input_type == 'positive_int':
            if type(input_value) != int:
                input_is_good = False
                print(input_name + " should be an integer!")
            else:
                if input_value <= 0:
                    input_is_good = False
                    print(input_name + " should be a positive integer!")
        else:
            print("Unhandled input type: " + input_type)
            
        return input_is_good


    def get_a_list_of_games(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return

        print("Getting a list of games for run ID", self.run_id, "...")
        for league_to_scrape in self.leagues_to_scrape:
            existence_check = self.db_cursor.execute("SELECT EXISTS (SELECT 1 FROM ScrapedRuns WHERE run_id=? AND league_url=?)", (self.run_id, league_to_scrape["url"]) ).fetchone()[0]
            if existence_check == 1:
                continue # already was scraped

            # if here, must scrape
            try:
                self.destroy_handles_and_create_new_one()
                self.driver.get(league_to_scrape["url"])
                wait_for_games = self.wait_by_xpath("//a[@href]/div[contains(@class, 'gameInfo')]/..", 30)
                if wait_for_games == 0:
                    continue

                innerHTML = self.driver.execute_script("return document.body.innerHTML")
                htmlElem = html.document_fromstring(innerHTML)
                game_links = []
                for game_el in htmlElem.xpath("//a[@href]/div[contains(@class, 'gameInfo')]/.."):
                    game_url_to_add = urljoin("https://www.pinnacle.com/", game_el.attrib["href"])
                    if game_url_to_add.endswith("/"):
                        game_url_to_add = game_url_to_add[0:-1]

                    game_links.append(game_url_to_add)
                        
                if len(game_links) == 0:
                    print("Couldn't find any games at", league_to_scrape["url"])
                    continue
                else:
                    for game_to_insert in game_links:
                        self.db_cursor.execute("INSERT OR IGNORE INTO Matches(game_url, run_id, league_name, league_url) VALUES(?,?,?,?)",
                                               (game_to_insert, self.run_id, league_to_scrape["league"], league_to_scrape["url"]))
                    self.db_cursor.execute("INSERT INTO ScrapedRuns(run_id, league_url, number_of_games) VALUES(?,?,?)", (self.run_id, league_to_scrape["url"], len(game_links) ))
                    self.db_conn.commit()
                    print("Found", len(game_links), "games for", league_to_scrape["league"])
                
            except KeyboardInterrupt:
                print("Manual interrupt, quit!")
                self.is_interrupted = True
                return
            except Exception as exc:
                print("An exception at", league_to_scrape, ":", repr(exc))
                continue
        return


    def scrape_odds(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return

        print("Scraping unscraped matches...")
        items_to_scrape = self.db_cursor.execute("SELECT game_url, run_id FROM Matches WHERE odds_html IS NULL AND run_id=?", (self.run_id,)).fetchall()
        print("Matches left to scrape:", len(items_to_scrape))

        for item_to_scrape in items_to_scrape:
            try:
                # check if maybe needed to skip due to future match!
                this_match_timestamp = self.db_cursor.execute("SELECT match_timestamp FROM Matches WHERE game_url=? AND match_timestamp IS NOT NULL", (item_to_scrape[0],)).fetchone()
                if this_match_timestamp != None:
                    if this_match_timestamp[0] - time.time() >= self.max_future_days*24*3600:
                        print("Skip", item_to_scrape[0], "since it is too much in the future!")
                        continue # skip this match because it is too much in the future!
                
                # if still here, have to scrape!
                self.destroy_handles_and_create_new_one()
                self.driver.get(item_to_scrape[0] + "/#period:0")
                wait_for_data = self.wait_by_xpath("//div[@data-test-id='Matchup Header']//div[contains(@class, 'startTime')]", 30)
                if wait_for_data == 0:
                    continue

                innerHTML = self.driver.execute_script("return document.body.innerHTML") # save if good
                htmlElem = html.document_fromstring(innerHTML)
                this_match = {"home":None, "away":None, "start_timetext":None, "start_timestamp":None}

                team_els = htmlElem.xpath("//div[@data-test-id='Matchup Header']//label[contains(@class, 'participantName')]")
                if len(team_els) == 2:
                    this_match["home"] = team_els[0].text_content().strip()
                    this_match["away"] = team_els[1].text_content().strip()

                starttime_el = htmlElem.xpath("//div[@data-test-id='Matchup Header']//div[contains(@class, 'startTime')]")
                if len(starttime_el) != 0:
                    try:
                        start_time_full = starttime_el[0].text_content().strip()
                        start_time_object = datetime.strptime(start_time_full[start_time_full.find(",")+1 : ].replace(" at ", " ").strip(), "%B %d, %Y %H:%M")
                        this_match["start_timetext"] = start_time_object.strftime("%Y-%m-%d %H:%M")
                        this_match["start_timestamp"] = start_time_object.timestamp()
                    except (ValueError, TypeError):
                        pass

                # see if good, update if needed!
                if this_match["home"] != None and this_match["away"] != None and this_match["start_timestamp"] != None:
                    # good to save
                    current_time_object = datetime.now()
                    if this_match["start_timestamp"] - time.time() >= self.max_future_days*24*3600:
                        data_to_save = None
                    else:
                        data_to_save = bz2.compress(pickle.dumps(innerHTML))

                    self.db_cursor.execute("UPDATE Matches SET home_team=?, away_team=?, match_time=?, match_timestamp=?, odds_html=?, time_of_scraping=?, time_of_scraping_stamp=? WHERE game_url=? AND run_id=?",
                                           (this_match["home"], this_match["away"], this_match["start_timetext"], this_match["start_timestamp"], data_to_save, current_time_object.strftime("%Y-%m-%d %H:%M"),
                                            current_time_object.timestamp(), item_to_scrape[0], item_to_scrape[1] ))
                    self.db_conn.commit()
                    print("Scraped match", items_to_scrape.index(item_to_scrape)+1, "/", len(items_to_scrape))
                else:
                    print("Couldn't find all data at", item_to_scrape[0])
                
            except KeyboardInterrupt:
                print("Manual interrupt, quit!")
                self.is_interrupted = True
                return
            except:
                print("An exception at", item_to_scrape)
                continue
        return


    def wait_by_xpath(self, xp, how_long_to_wait): # xp is string, how_long_to_wait float - the number of seconds to wait
        try:
            WebDriverWait(self.driver, how_long_to_wait).until(EC.presence_of_element_located((By.XPATH, xp)) )
            time.sleep(self.TIME_PAUSE)
            return 1 # success
        except TimeoutException:
            print ("Too much time has passed while waiting for", xp)
            return 0 # fail

            
    def fix_string(self, entry_string): # remove "\n", "\t" and double spaces
        exit_string = entry_string.replace("\n", "")
        exit_string = exit_string.replace("\t", "")
        exit_string = exit_string.replace("\r", "")
        while "  " in exit_string:
            exit_string = exit_string.replace("  ", " ")
        if len(exit_string) > 0: # remove first space
            if exit_string[0] == ' ':
                exit_string = exit_string[1:len(exit_string)]
        if len(exit_string) > 0: # remove last space
            if exit_string[len(exit_string)-1] == ' ':
                exit_string = exit_string[0:len(exit_string)-1]

        return exit_string


    def start_driver_normal(self):
        #service = Service(self.PATH_TO_BROWSER)
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-search-engine-choice-screen")
        
        #normal_driver = webdriver.Chrome(service=service, options=options)
        normal_driver = webdriver.Chrome(options=options)
        normal_driver.maximize_window()
        return normal_driver


    def start_driver_headless(self):
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("window-size=1920,1080")
        return webdriver.Chrome(options=chrome_options)


    def destroy_handles_and_create_new_one(self):
        # call it before opening an url
        while 1:
            initial_handles = self.driver.window_handles
            self.driver.execute_script("window.open();")
            handles_after_opening = self.driver.window_handles
            if len(handles_after_opening) > len(initial_handles):
                break
            else:
                print("Couldn't open a handle!")
                time.sleep(10.0)
                continue
            
        added_handle = []
        for handle in handles_after_opening:
            if handle in initial_handles:
                self.driver.switch_to.window(handle)
                self.driver.close()
            else:
                added_handle.append(handle)

        self.driver.switch_to.window(added_handle[0])
        return





if __name__ == '__main__':
    scraper = PinnacleScraper(RUN_ID, INPUT_FILE, INPUT_SHEET, DATABASE_NAME, PATH_TO_BROWSER, INPUT_MAX_FUTURE_DAYS, mode='headless')
    scraper.get_a_list_of_games()
    scraper.scrape_odds()
    if hasattr(scraper, 'driver'):
        scraper.driver.quit()
