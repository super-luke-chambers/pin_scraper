WRITE_ONLY_LAST_RUN_ID = True # True to get recent data, False to get a bit more in the past
MAX_PAST_DAYS_TO_WRITE = 7 # to get more history, only relevant if ^^ is False !

from pinnacle_scrape_odds_1_web import RUN_ID
from pinnacle_scrape_odds_1_web import INPUT_FILE
from pinnacle_scrape_odds_1_web import INPUT_SHEET
from pinnacle_scrape_odds_1_web import DATABASE_NAME
from pinnacle_scrape_odds_1_web import PATH_TO_BROWSER
from pinnacle_scrape_odds_1_web import INPUT_MAX_FUTURE_DAYS
from pinnacle_scrape_odds_1_web import PinnacleScraper

# install: requests, lxml, openpyxl
import requests
from lxml import html
from openpyxl import Workbook, load_workbook
import os
import sys
import bz2
import json
import pickle
import pprint
import threading
import time
from datetime import datetime, timedelta
import sqlite3
import re


class PinnacleWriter(PinnacleScraper):
    def __init__(self, write_only_last_run_id, max_past_days_to_write, *args):
        super().__init__(*args, mode='write')

        # check inputs!
        if type(write_only_last_run_id) != bool:
            print("WRITE_ONLY_LAST_RUN_ID must be True or False")
            self.inputs_are_good = False

        if self.check_input('MAX_PAST_DAYS_TO_WRITE', 'positive_int', max_past_days_to_write) == False:
            self.inputs_are_good = False

        if self.inputs_are_good == False:
            return
        
        # set inputs!
        self.write_only_last_run_id = write_only_last_run_id
        self.max_past_days_to_write = max_past_days_to_write
        return
    
    def write_data(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return
        
        print("Writing data...")
        HEADERS = ['Game URL', 'Match Start Time', 'League', 'Home', 'Away', 'Time of scraping', 'Home Win', 'Draw', 'Away Win', 'AH Line', 'AH Home Odds', 'AH Away Odds',
                   'Asian Corners Line', 'Asian Corners O Odds', 'Asian Corners U Odds', 'Goal Line', 'Goal O Odds', 'Goal U Odds',
                   'Corner Handicap', 'Home Handicap Corners Odds', 'Away Handicap Corners Odds', 'Bookings', 'Bookings O Odds', 'Bookings U Odds',
                   'Bookings Handicap', 'Bookings Home Handicap Odds', 'Bookings Away Handicap Odds']
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.title = 'Sheet1'
        ws.append(HEADERS)

        # form a query!
        if self.write_only_last_run_id == True:
            fetcher = self.db_cursor.execute("SELECT * FROM Matches WHERE run_id=? AND odds_html IS NOT NULL ORDER BY match_timestamp ASC", (self.run_id,)) # remove limit!
        else:
            fetcher = self.db_cursor.execute("SELECT * FROM Matches WHERE odds_html IS NOT NULL AND time_of_scraping_stamp >=? ORDER BY time_of_scraping_stamp DESC",
                                             (time.time()-self.max_past_days_to_write*24*3600,)) # remove limit!

        for fetched_row in fetcher:
            parsed_match_data = self.parse_data(fetched_row[0], fetched_row[2], fetched_row[3], fetched_row[4], fetched_row[6], fetched_row[8], fetched_row[9])
            #pprint.pprint(parsed_match_data)
            if parsed_match_data["Home Win"] == '':
                continue
            
            row_to_write = [parsed_match_data[header_item] for header_item in HEADERS]
            ws.append(row_to_write)


        outfile_name = "pinnacle_odds_data.xlsx"
        wb.save(outfile_name)
        print("Created output file:", outfile_name)
        return


    def parse_data(self, input_gameurl, input_hometeam, input_awayteam, input_matchtime, input_league, input_oddshtml, input_scrapetime):
        data_to_return = {}
        data_to_return["Game URL"] = input_gameurl
        data_to_return["Match Start Time"] = input_matchtime
        data_to_return["League"] = input_league
        data_to_return["Home"] = input_hometeam
        data_to_return["Away"] = input_awayteam
        data_to_return["Time of scraping"] = input_scrapetime

        tree = html.document_fromstring(pickle.loads(bz2.decompress(input_oddshtml)))
        
        # parse out from html!
        data_to_return["Home Win"] = ''
        data_to_return["Draw"] = ''
        data_to_return["Away Win"] = ''
        moneyline_btn_els = tree.xpath("//div/span[text()='Money Line – Match']/../following-sibling::div[1]/div[contains(@class, 'buttons')]/div[contains(@class, 'buttonRow')]//div[contains(@class, 'buttonWrapper')]/button")
        for moneyline_index, moneyline_btn_el in enumerate(moneyline_btn_els):
            moneyline_text = ""
            moneyline_odds = ""
            
            moneyline_label_el = moneyline_btn_el.xpath("./span[contains(@class, 'label')]")
            if len(moneyline_label_el) != 0:
                moneyline_text = self.fix_string(moneyline_label_el[0].text_content())

            moneyline_odds_el = moneyline_btn_el.xpath("./span[contains(@class, 'price')]")
            if len(moneyline_odds_el) != 0:
                moneyline_odds = self.fix_string(moneyline_odds_el[0].text_content())

            if moneyline_text.lower() == 'draw':
                data_to_return["Draw"] = moneyline_odds
            else:
                if moneyline_index == 0:
                    data_to_return["Home Win"] = moneyline_odds
                elif moneyline_index in [1,2]:
                    data_to_return["Away Win"] = moneyline_odds
                else:
                    pass
                    

        
        data_to_return["AH Line"] = ''
        data_to_return["AH Home Odds"] = ''
        data_to_return["AH Away Odds"] = ''
        ah_row_els = tree.xpath("//div/span[text()='Handicap – Match']/../following-sibling::div[1]/div[contains(@class, 'buttons')]/div[contains(@class, 'buttonRow')]")
        if len(ah_row_els) != 0:
            all_ah_objects = [self.get_line_and_odds(ah_row_el, 'home/away') for ah_row_el in ah_row_els]
            ah_object = self.get_most_even_odds_object(all_ah_objects)
            data_to_return["AH Line"] = ah_object["line"]
            data_to_return["AH Home Odds"] = ah_object["odds_home"]
            data_to_return["AH Away Odds"] = ah_object["odds_away"]
            

        data_to_return["Asian Corners Line"] = ''
        data_to_return["Asian Corners O Odds"] = ''
        data_to_return["Asian Corners U Odds"] = ''
        ascorner_row_els = tree.xpath("//div/span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'total') and contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'corners') and contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'match')]/../following-sibling::div[1]/div[contains(@class, 'buttons')]/div[contains(@class, 'buttonRow')]")
        if len(ascorner_row_els) != 0:
            all_asian_corners_objects = [self.get_line_and_odds(ascorner_row_el, 'over/under') for ascorner_row_el in ascorner_row_els]
            asian_corners_object = self.get_most_even_odds_object(all_asian_corners_objects)
            data_to_return["Asian Corners Line"] = asian_corners_object["line"]
            data_to_return["Asian Corners O Odds"] = asian_corners_object["odds_over"]
            data_to_return["Asian Corners U Odds"] = asian_corners_object["odds_under"]


        data_to_return["Goal Line"] = ''
        data_to_return["Goal O Odds"] = ''
        data_to_return["Goal U Odds"] = ''
        goaloverunder_row_els = tree.xpath("//div/span[text()='Total – Match']/../following-sibling::div[1]/div[contains(@class, 'buttons')]/div[contains(@class, 'buttonRow')]")
        if len(goaloverunder_row_els) != 0:
            all_goals_objects = [self.get_line_and_odds(goaloverunder_row_el, 'over/under') for goaloverunder_row_el in goaloverunder_row_els]
            goals_object = self.get_most_even_odds_object(all_goals_objects)
            data_to_return["Goal Line"] = goals_object["line"]
            data_to_return["Goal O Odds"] = goals_object["odds_over"]
            data_to_return["Goal U Odds"] = goals_object["odds_under"]


        data_to_return["Corner Handicap"] = ''
        data_to_return["Home Handicap Corners Odds"] = ''
        data_to_return["Away Handicap Corners Odds"] = ''
        corner_handi_row_els = tree.xpath("//div/span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'handicap') and contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'corners') and contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'match')]/../following-sibling::div[1]/div[contains(@class, 'buttons')]/div[contains(@class, 'buttonRow')]")
        if len(corner_handi_row_els) != 0:
            all_corner_handi_objects = [self.get_line_and_odds(corner_handi_row_el, 'home/away') for corner_handi_row_el in corner_handi_row_els]
            corner_handi_object = self.get_most_even_odds_object(all_corner_handi_objects)
            data_to_return["Corner Handicap"] = corner_handi_object["line"]
            data_to_return["Home Handicap Corners Odds"] = corner_handi_object["odds_home"]
            data_to_return["Away Handicap Corners Odds"] = corner_handi_object["odds_away"]


        data_to_return["Bookings"] = ''
        data_to_return["Bookings O Odds"] = ''
        data_to_return["Bookings U Odds"] = ''
        bookings_over_under_row_els = tree.xpath("//div/span[text()='Total (Bookings) – Match']/../following-sibling::div[1]/div[contains(@class, 'buttons')]/div[contains(@class, 'buttonRow')]")
        if len(bookings_over_under_row_els) != 0:
            all_bookings_objects = [self.get_line_and_odds(bookings_over_under_row_el, 'over/under') for bookings_over_under_row_el in bookings_over_under_row_els]
            bookings_object = self.get_most_even_odds_object(all_bookings_objects)
            data_to_return["Bookings"] = bookings_object["line"]
            data_to_return["Bookings O Odds"] = bookings_object["odds_over"]
            data_to_return["Bookings U Odds"] = bookings_object["odds_under"]


        data_to_return["Bookings Handicap"] = ''
        data_to_return["Bookings Home Handicap Odds"] = ''
        data_to_return["Bookings Away Handicap Odds"] = ''
        bookings_handi_row_els = tree.xpath("//div/span[text()='Handicap (Bookings) – Match']/../following-sibling::div[1]/div[contains(@class, 'buttons')]/div[contains(@class, 'buttonRow')]")
        if len(bookings_handi_row_els) != 0:
            all_bookings_handi_objects = [self.get_line_and_odds(bookings_handi_row_el, 'home/away') for bookings_handi_row_el in bookings_handi_row_els]
            bookings_handi_object = self.get_most_even_odds_object(all_bookings_handi_objects)
            data_to_return["Bookings Handicap"] = bookings_handi_object["line"]
            data_to_return["Bookings Home Handicap Odds"] = bookings_handi_object["odds_home"]
            data_to_return["Bookings Away Handicap Odds"] = bookings_handi_object["odds_away"]
            
        return data_to_return


    def get_line_and_odds(self, input_element_row, input_mode):
        data_to_return = {"line":"", "odds_over":"", "odds_under":"", "odds_home":"", "odds_away":""}

        btn_els = input_element_row.xpath(".//div[contains(@class, 'buttonWrapper')]/button")
        if len(btn_els) == 2:
            # find line first
            line_el = btn_els[0].xpath("./span[contains(@class, 'label')]")
            if len(line_el) != 0:
                line_match = re.findall("[-+0-9.]+", self.fix_string(line_el[0].text_content()))
                if len(line_match) != 0:
                    data_to_return["line"] = line_match[0]

            # get first odd and other odd.
            for btn_index, btn_el in enumerate(btn_els):
                label_text = ""
                odds_text = ""
                label_el = btn_el.xpath("./span[contains(@class, 'label')]")
                if len(label_el) != 0:
                    label_text = self.fix_string(label_el[0].text_content().lower())

                odds_el = btn_el.xpath("./span[contains(@class, 'price')]")
                if len(odds_el) != 0:
                    odds_text = self.fix_string(odds_el[0].text_content())

                if input_mode == 'over/under':
                    if 'over' in label_text:
                        data_to_return["odds_over"] = odds_text
                    elif 'under' in label_text:
                        data_to_return["odds_under"] = odds_text
                    else:
                        pass

                elif input_mode == 'home/away':
                    if btn_index == 0:
                        data_to_return["odds_home"] = odds_text
                    elif btn_index == 1:
                        data_to_return["odds_away"] = odds_text
                    else:
                        pass
                else:
                    pass
                
        return data_to_return


    def get_most_even_odds_object(self, list_of_dict_odds):
        index_to_return = 0
        min_diff = None
        for dict_index, odds_dict in enumerate(list_of_dict_odds):
            odds_to_consider = []
            for key_to_check in ["odds_over", "odds_under", "odds_home", "odds_away"]:
                try:
                    cur_odds = float(odds_dict[key_to_check])
                    odds_to_consider.append(cur_odds)
                except (ValueError, TypeError, ZeroDivisionError):
                    continue

            if len(odds_to_consider) == 2:
                cur_diff = abs(odds_to_consider[0] - odds_to_consider[1])
                if min_diff == None:
                    min_diff = cur_diff
                    index_to_return = dict_index
                else:
                    if cur_diff < min_diff:
                        min_diff = cur_diff
                        index_to_return = dict_index
                        
        
        return list_of_dict_odds[index_to_return]

    

    def convert_odds_from_US_to_decimal(self, input_us_number):
        value_to_return = None
        if type(input_us_number) in [int, float]:
            value_to_convert = float(input_us_number)
            if value_to_convert > 0:
                value_to_return = value_to_convert/100 + 1.0
            elif value_to_convert < 0:
                value_to_return = 1.0 - 100/value_to_convert
            
        return value_to_return

if __name__ == '__main__':
    pars = PinnacleWriter(WRITE_ONLY_LAST_RUN_ID, MAX_PAST_DAYS_TO_WRITE, RUN_ID, INPUT_FILE, INPUT_SHEET, DATABASE_NAME, PATH_TO_BROWSER, INPUT_MAX_FUTURE_DAYS)
    pars.write_data()
